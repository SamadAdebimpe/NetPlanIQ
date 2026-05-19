import datetime
from rich import print
from netaddr import IPAddress
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from ip_planner.core import get_remaining_cidr_blocks

# Sanitize worksheet name
def sanitize_sheet_name(name):
    invalid_characters = ["\\", "/", "?", "*", "[", "]", ":"]

    cleaned_name = str(name).strip()

    for character in invalid_characters:
        cleaned_name = cleaned_name.replace(character, "_")

    if not cleaned_name:
        cleaned_name = "Sheet"

    return cleaned_name[:31]

# GET UNIQUE WORKSHEET NAME (NO DUPLICATE)
def get_unique_sheet_name(workbook, base_name):
    sheet_name = sanitize_sheet_name(base_name)

    if sheet_name not in workbook.sheetnames:
        return sheet_name

    counter = 1

    while True:
        suffix = f"_{counter}"
        max_name_length = 31 - len(suffix)
        candidate_name = f"{sheet_name[:max_name_length]}{suffix}"

        if candidate_name not in workbook.sheetnames:
            return candidate_name

        counter += 1


# SAFE FILENAME 
def safe_filename_part(value):
    invalid_characters = ["\\", "/", ":", "*", "?", '"', "<", ">", "|"]

    cleaned_value = str(value).strip()

    for character in invalid_characters:
        cleaned_value = cleaned_value.replace(character, "_")

    return cleaned_value

# FORMAT THE WORKSHEET
def format_worksheet(sheet):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(bold=True, size=12)

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    if sheet.title == "Plan Summary":
        for cell in sheet["A"]:
            if cell.value:
                cell.font = title_font
        sheet.freeze_panes = "A2"

    else:
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill

        sheet.freeze_panes = "A2"

        if sheet.max_row > 1 and sheet.max_column > 1:
            sheet.auto_filter.ref = sheet.dimensions

    for column_cells in sheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        adjusted_width = min(max_length + 2, 45)
        sheet.column_dimensions[column_letter].width = adjusted_width

# Export host based to excel
def export_to_excel(planned_subnets, base_network=None, next_free_ip=None, filename=None, include_ip_sheets=True):
    try:
        if filename is None:
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

            if base_network is not None:
                network_part = safe_filename_part(str(base_network))
                filename = f"ip_plan_{network_part}_{timestamp}.xlsx"
            else:
                filename = f"ip_plan_{timestamp}.xlsx"

        workbook = Workbook()

        summary_sheet = workbook.active
        summary_sheet.title = "Plan Summary"

        if base_network is not None:
            total_available = base_network.size

            if next_free_ip is not None:
                total_consumed = next_free_ip - int(base_network.network)
            else:
                total_consumed = sum(subnet["total_addresses"] for subnet in planned_subnets)

            remaining_addresses = total_available - total_consumed

            summary_sheet.append(["Plan Summary", ""])
            summary_sheet.append(["Base Network", str(base_network)])
            summary_sheet.append(["Network Address", str(base_network.network)])
            summary_sheet.append(["CIDR Prefix", f"/{base_network.prefixlen}"])
            summary_sheet.append(["Subnet Mask", str(base_network.netmask)])
            summary_sheet.append(["Wildcard", str(base_network.hostmask)])
            summary_sheet.append(["Broadcast Address", str(base_network.broadcast)])
            summary_sheet.append(["Total Addresses Available", total_available])
            summary_sheet.append(["Total Address Space Consumed", total_consumed])
            summary_sheet.append(["Remaining Addresses", remaining_addresses])
            summary_sheet.append(["Generated On", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        else:
            summary_sheet.append(["Plan Summary", ""])
            summary_sheet.append(["Generated On", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

        allocation_sheet = workbook.create_sheet(title="Allocations")

        allocation_sheet.append([
            "Name",
            "Type",
            "Hosts Needed",
            "Network Address",
            "Prefix",
            "Subnet Mask",
            "Wildcard",
            "Broadcast Address",
            "Total Addresses",
            "Usable Hosts",
            "First Usable IP",
            "Last Usable IP",
            "Suggested Gateway",
            "Unused Hosts",
            "Preference",
            "Adjusted Hosts",
            "Efficiency %",
            "Recommendation",
            "Explanation",
        ])

        for subnet_info in planned_subnets:
            allocation_sheet.append([
                subnet_info["name"],
                subnet_info["type"],
                subnet_info["hosts_needed"],
                str(subnet_info["network"]),
                subnet_info["prefix"],
                str(subnet_info["subnet_mask"]),
                str(subnet_info["wildcard"]),
                str(subnet_info["broadcast"]),
                subnet_info["total_addresses"],
                subnet_info["usable_hosts"],
                str(subnet_info["first_ip"]),
                str(subnet_info["last_ip"]),
                str(subnet_info["suggested_gateway"]),
                subnet_info["unused_hosts"],
                subnet_info["preference"],
                subnet_info["adjusted_hosts"],
                subnet_info["efficiency"],
                subnet_info["recommendation"],
                subnet_info["explanation"],
            ])

        if base_network is not None and next_free_ip is not None:
            remaining_sheet = workbook.create_sheet(title="Remaining Blocks")

            remaining_sheet.append([
                "CIDR Block",
                "Network Address",
                "Prefix",
                "Subnet Mask",
                "Wildcard",
                "Broadcast Address",
                "Total Addresses",
                "Usable Hosts",
                "First Usable IP",
                "Last Usable IP",
                "Note",
            ])

            remaining_blocks = get_remaining_cidr_blocks(base_network, next_free_ip)

            if not remaining_blocks:
                remaining_sheet.append([
                    "N/A",
                    "",
                    "",
                    "",
                    "",
                    "",
                    0,
                    0,
                    "",
                    "",
                    "No unallocated address space remains.",
                ])
            else:
                for block in remaining_blocks:
                    usable_hosts = block.size - 2 if block.prefixlen < 31 else block.size
                    first_ip = block[1] if block.prefixlen < 31 else block[0]
                    last_ip = block[-2] if block.prefixlen < 31 else block[-1]

                    remaining_sheet.append([
                        str(block),
                        str(block.network),
                        block.prefixlen,
                        str(block.netmask),
                        str(block.hostmask),
                        str(block.broadcast),
                        block.size,
                        usable_hosts,
                        str(first_ip),
                        str(last_ip),
                        "Available for future allocation",
                    ])

        if include_ip_sheets:
            for subnet_info in planned_subnets:
                sheet_name = get_unique_sheet_name(workbook, subnet_info["name"])
                dept_sheet = workbook.create_sheet(title=sheet_name)

                dept_sheet.append(["IP Address / Range", "Status", "Default Gateway", "Note"])

                if subnet_info["first_ip"] == "N/A" or subnet_info["last_ip"] == "N/A":
                    dept_sheet.append([
                        str(subnet_info["network"]),
                        "Network",
                        "",
                        "No traditional usable host range available for this subnet.",
                    ])
                    continue

                if subnet_info["usable_hosts"] > 3000:
                    dept_sheet.append([
                        str(subnet_info["network"]),
                        "Network Address",
                        "",
                        "Do not assign to a device",
                    ])

                    dept_sheet.append([
                        str(subnet_info["suggested_gateway"]),
                        "Default Gateway",
                        str(subnet_info["suggested_gateway"]),
                        "Assign to router, firewall, or Layer 3 switch interface",
                    ])

                    available_start = IPAddress(int(subnet_info["first_ip"]) + 1)

                    dept_sheet.append([
                        f"{available_start} - {subnet_info['last_ip']}",
                        "Available Range",
                        str(subnet_info["suggested_gateway"]),
                        "Usable for hosts, servers, printers, APs, or DHCP pool",
                    ])

                    dept_sheet.append([
                        str(subnet_info["broadcast"]),
                        "Broadcast Address",
                        "",
                        "Do not assign to a device",
                    ])

                    dept_sheet.append([
                        str(subnet_info["usable_hosts"]),
                        "Usable Hosts",
                        "",
                        "Total assignable host addresses in this subnet",
                    ])

                else:
                    dept_sheet.append([
                        str(subnet_info["network"]),
                        "Network Address",
                        "",
                        "Do not assign to a device",
                    ])

                    first_ip = IPAddress(subnet_info["first_ip"])
                    last_ip = IPAddress(subnet_info["last_ip"])
                    gateway = IPAddress(subnet_info["suggested_gateway"])

                    current_ip = int(first_ip)

                    while current_ip <= int(last_ip):
                        ip_address = IPAddress(current_ip)

                        if ip_address == gateway:
                            status = "Default Gateway"
                            note = "Assign to router, firewall, or Layer 3 switch interface"
                        else:
                            status = "Available"
                            note = "Available for host assignment"

                        dept_sheet.append([
                            str(ip_address),
                            status,
                            str(gateway),
                            note,
                        ])

                        current_ip += 1

                    dept_sheet.append([
                        str(subnet_info["broadcast"]),
                        "Broadcast Address",
                        "",
                        "Do not assign to a device",
                    ])

        for sheet in workbook.worksheets:
            format_worksheet(sheet)

        workbook.save(filename)
        print(f"\n[bold green]Excel Export completed: {filename}[/bold green]")

    except Exception as e:
        print(f"[bold red]Excel export failed: {e}[/bold red]")

# 
# Export P2P NETWORK to excel
def export_p2p_to_excel(network, filename=None):
    try:
        if filename is None:
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            network_part = safe_filename_part(str(network))
            filename = f"p2p_plan_{network_part}_{timestamp}.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Point-to-Point"

        sheet.append(["Field", "Value"])
        sheet.append(["Network Address", str(network.network)])
        sheet.append(["CIDR Prefix", f"/{network.prefixlen}"])
        sheet.append(["Subnet Mask", str(network.netmask)])
        sheet.append(["Wildcard", str(network.hostmask)])
        sheet.append(["Broadcast Address", str(network.broadcast)])
        sheet.append(["Total Addresses", network.size])

        if network.prefixlen == 31:
            sheet.append(["Usable IP 1", str(network[0])])
            sheet.append(["Usable IP 2", str(network[1])])
            sheet.append(["Usable Hosts", 2])
            sheet.append(["Note", "/31 is commonly used for point-to-point links under RFC 3021."])
        elif network.prefixlen == 30:
            sheet.append(["First Usable IP", str(network[1])])
            sheet.append(["Last Usable IP", str(network[-2])])
            sheet.append(["Usable Hosts", 2])
            sheet.append(["Note", "/30 provides two usable host addresses for point-to-point links."])

        format_worksheet(sheet)

        workbook.save(filename)
        print(f"\n[bold green]Point-to-point Excel export completed: {filename}[/bold green]")

    except Exception as e:
        print(f"[bold red]Point-to-point Excel export failed: {e}[/bold red]")
